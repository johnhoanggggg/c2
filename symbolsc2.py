import fitz  # PyMuPDF
import cv2
import numpy as np
import io
import math
import re
import os
import time
import random
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Dict, Any, Optional, Pattern
from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, FancyArrowPatch


# ─────────────────────────────────────────────────────────────────────────────
# 1. Configuration & Constants
# ─────────────────────────────────────────────────────────────────────────────
# REMOVED: PIXEL_2_MM_SCALING — no longer needed.
# Coordinates are now exported in native PDF points (1 pt = 1/72 inch),
# matching the coordinate system of the original PDF document.

@dataclass
class PrecomputedTemplate:
    image: np.ndarray
    mask: np.ndarray
    rotation: int
    flipped: bool
    w: int
    h: int

@dataclass
class DetectionConfig:
    template_path: str
    export_macro: str
    threshold: float = 0.5
    overlap_threshold: float = 0.5
    search_radius: int = 100
    rotate_vary: bool = False 
    flip_vary: bool = False 
    label_required: bool = True
    use_mask: bool = True
    num_terminals: int = 1
    
    function_regex: Optional[str] = None
    mark_regex: Optional[str] = None
    location_regex: Optional[str] = r"^\+[A-Z].*"
    terminal_number_regex: Optional[str] = r"^\d{1,2}$"
    terminal_child_regex: Optional[str] = r"^\d{1,2}([A-Za-z])?$"

    _re_function: Optional[Pattern] = field(init=False, default=None)
    _re_mark: Optional[Pattern] = field(init=False, default=None)
    _re_location: Optional[Pattern] = field(init=False, default=None)
    _re_term_num: Optional[Pattern] = field(init=False, default=None)
    _re_term_child: Optional[Pattern] = field(init=False, default=None)
    
    _precomputed_variants: List[PrecomputedTemplate] = field(init=False, default_factory=list)

    def __post_init__(self):
        self._re_function = re.compile(self.function_regex) if self.function_regex else None
        self._re_mark = re.compile(self.mark_regex) if self.mark_regex else None
        self._re_location = re.compile(self.location_regex) if self.location_regex else None
        self._re_term_num = re.compile(self.terminal_number_regex) if self.terminal_number_regex else None
        self._re_term_child = re.compile(self.terminal_child_regex) if self.terminal_child_regex else None

    def precompute_templates(self):
        if not Path(self.template_path).exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        templ_bgr = cv2.imread(str(self.template_path))
        if templ_bgr is None:
            raise ValueError(f"Could not load image: {self.template_path}")
            
        templ_gray = cv2.cvtColor(templ_bgr, cv2.COLOR_BGR2GRAY)
        base_mask = np.where(templ_gray < 150, 255, 0).astype(np.uint8)

        rotations = [0, 90, 180, 270] if self.rotate_vary else [0]
        self._precomputed_variants = []

        for rot in rotations:
            t_curr = self._rotate_image(templ_gray, rot)
            m_curr = self._rotate_image(base_mask, rot)

            variants = [("normal", t_curr, m_curr)]
            if self.flip_vary:
                variants.append(("flipped", cv2.flip(t_curr, 1), cv2.flip(m_curr, 1)))
            
            for v_name, t, m in variants:
                h, w = t.shape
                final_mask = m if self.use_mask else None
                self._precomputed_variants.append(PrecomputedTemplate(
                    image=t, mask=final_mask, rotation=rot, flipped=(v_name == "flipped"), w=w, h=h
                ))
    
    @staticmethod
    def _rotate_image(image, angle):
        if angle == 0: return image
        if angle == 90: return cv2.rotate(image, cv2.ROTATE_90_COUNTERCLOCKWISE)
        if angle == 180: return cv2.rotate(image, cv2.ROTATE_180)
        if angle == 270: return cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
        return image

# ─────────────────────────────────────────────────────────────────────────────
# 2. Data Models & Spatial Index
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class TextWord:
    text: str
    rect: Tuple[float, float, float, float]
    center_scaled: Tuple[float, float] = (0.0, 0.0)
    original_index: int = -1
    angle: float = 0.0

@dataclass
class SymbolData:
    template_name: str
    export_macro: str
    center: Tuple[int, int]
    bbox: Tuple[int, int, int, int]
    score: float
    rotation: int
    page_num: int
    flipped: bool = False
    function: str = "\u00A0"
    mark: str = ""
    location: str = ""
    terminals_found: Dict[str, str] = field(default_factory=dict)
    
    terminals_exact_coords: Dict[str, Tuple[int, int, int]] = field(default_factory=dict)
    
    terminal_parent: str = "\u00A0"
    terminal_parent_point: Optional[Tuple[int, int]] = None
    terminal_parent_angle: int = 0
    
    is_inferred: bool = False
    assigned_header: Optional['SymbolData'] = field(default=None, repr=False)

class SpatialTextIndex:
    def __init__(self, words: List[TextWord], cell_size: int = 150):
        self.cell_size = cell_size
        self.grid = defaultdict(list)
        self.words = words
        for idx, w in enumerate(words):
            w.original_index = idx
            gx = int(w.center_scaled[0] // cell_size)
            gy = int(w.center_scaled[1] // cell_size)
            self.grid[(gx, gy)].append(w)

    def query_nearby(self, center: Tuple[float, float], radius: float) -> List[TextWord]:
        cx, cy = center
        min_gx = int((cx - radius) // self.cell_size)
        max_gx = int((cx + radius) // self.cell_size)
        min_gy = int((cy - radius) // self.cell_size)
        max_gy = int((cy + radius) // self.cell_size)

        candidates = []
        for gx in range(min_gx, max_gx + 1):
            for gy in range(min_gy, max_gy + 1):
                candidates.extend(self.grid.get((gx, gy), []))
        return candidates

# ─────────────────────────────────────────────────────────────────────────────
# 3. PDF Service
# ─────────────────────────────────────────────────────────────────────────────
class PDFService:
    @staticmethod
    def render_page(pdf_path: str, page_num: int, zoom: float) -> Tuple[np.ndarray, List[TextWord], Tuple[float, float]]:
        doc = fitz.open(pdf_path)
        page = doc.load_page(page_num)
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img_data = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
        if pix.n >= 3:
            image = cv2.cvtColor(img_data, cv2.COLOR_RGB2BGR)
        else:
            image = cv2.cvtColor(img_data, cv2.COLOR_GRAY2BGR)

        words = []
        raw = page.get_text("rawdict")
        for block in raw.get("blocks", []):
            if block.get("type") != 0: continue
            for line in block.get("lines", []):
                dx, dy = line.get("dir", (1.0, 0.0))
                angle_rad = math.atan2(dy, dx)
                angle_deg = angle_rad * 180.0 / math.pi
                for span in line.get("spans", []):
                    chars = span.get("chars", [])
                    if not chars: continue
                    current_word_chars = []
                    for char in chars:
                        c_val = char.get("c", "")
                        if c_val.strip() == "":
                            if current_word_chars:
                                PDFService._flush_word(current_word_chars, angle_deg, zoom, words)
                                current_word_chars = []
                        else:
                            current_word_chars.append(char)
                    if current_word_chars:
                        PDFService._flush_word(current_word_chars, angle_deg, zoom, words)
        dims = (page.rect.width, page.rect.height)
        doc.close()
        return image, words, dims

    @staticmethod
    def _flush_word(chars_list, angle, zoom, words_list):
        if not chars_list: return
        text = "".join([c["c"] for c in chars_list])
        xs0 = [c["bbox"][0] for c in chars_list]
        ys0 = [c["bbox"][1] for c in chars_list]
        xs1 = [c["bbox"][2] for c in chars_list]
        ys1 = [c["bbox"][3] for c in chars_list]
        x0, y0, x1, y1 = min(xs0), min(ys0), max(xs1), max(ys1)
        zx0, zy0, zx1, zy1 = x0 * zoom, y0 * zoom, x1 * zoom, y1 * zoom
        zcx = (zx0 + zx1) / 2
        zcy = (zy0 + zy1) / 2
        words_list.append(TextWord(text, (zx0, zy0, zx1, zy1), (zcx, zcy), angle=angle))

# ─────────────────────────────────────────────────────────────────────────────
# 4. Vision Service
# ─────────────────────────────────────────────────────────────────────────────
class VisionService:
    @staticmethod
    def match_precomputed(scene_gray: np.ndarray, variants: List[PrecomputedTemplate], threshold: float) -> List[Dict[str, Any]]:
        matches = []
        for v in variants:
            try:
                if v.mask is not None:
                    res = cv2.matchTemplate(scene_gray, v.image, cv2.TM_SQDIFF_NORMED, mask=v.mask)
                    res = 1.0 - res
                else:
                    res = cv2.matchTemplate(scene_gray, v.image, cv2.TM_CCOEFF_NORMED)
            except cv2.error:
                continue 
            
            ys, xs = np.where(res >= threshold)
            for x, y in zip(xs, ys):
                matches.append({
                    "bbox": (int(x), int(y), int(x + v.w), int(y + v.h)),
                    "center": (int(x + v.w // 2), int(y + v.h // 2)),
                    "score": float(res[y, x]),
                    "rotation": v.rotation,
                    "flipped": v.flipped
                })
        return VisionService._non_max_suppression(matches)

    @staticmethod
    def _non_max_suppression(candidates: List[Dict], iou_thresh: float = 0.3) -> List[Dict]:
        if not candidates: return []
        candidates = sorted(candidates, key=lambda x: x["score"], reverse=True)
        kept = []
        for c in candidates:
            if not any(VisionService._iou(c["bbox"], k["bbox"]) > iou_thresh for k in kept):
                kept.append(c)
        return kept

    @staticmethod
    def _iou(b1, b2):
        xA, yA = max(b1[0], b2[0]), max(b1[1], b2[1])
        xB, yB = min(b1[2], b2[2]), min(b1[3], b2[3])
        interArea = max(0, xB - xA) * max(0, yB - yA)
        boxAArea = (b1[2] - b1[0]) * (b1[3] - b1[1])
        boxBArea = (b2[2] - b2[0]) * (b2[3] - b2[1])
        denom = float(boxAArea + boxBArea - interArea)
        return interArea / denom if denom > 0 else 0

    @staticmethod
    def cross_class_nms(symbols_with_cfg: List[Tuple['SymbolData', 'DetectionConfig']], 
                        iou_threshold: float = 0.3) -> List[Tuple['SymbolData', 'DetectionConfig']]:
        if not symbols_with_cfg:
            return []
        
        sorted_symbols = sorted(symbols_with_cfg, key=lambda x: x[0].score, reverse=True)
        
        kept = []
        suppressed_count = 0
        
        for sym, cfg in sorted_symbols:
            cx, cy = sym.center
            is_suppressed = False
            
            for kept_sym, kept_cfg in kept:
                kx, ky = kept_sym.center
                
                iou = VisionService._iou(sym.bbox, kept_sym.bbox)
                
                if iou > iou_threshold:
                    is_suppressed = True
                    suppressed_count += 1
                    print(f"  [Cross-Class NMS] Suppressing '{sym.template_name}' (score={sym.score:.3f}) "
                          f"due to '{kept_sym.template_name}' (score={kept_sym.score:.3f}), "
                          f"iou={iou:.3f}")
                    break
            
            if not is_suppressed:
                kept.append((sym, cfg))
        
        if suppressed_count > 0:
            print(f"  [Cross-Class NMS] Suppressed {suppressed_count} lower-confidence detections")
        
        return kept

# ─────────────────────────────────────────────────────────────────────────────
# 5. Text Analysis Service
# ─────────────────────────────────────────────────────────────────────────────
class TextService:
    @staticmethod
    def assign_unique_labels(symbols_with_cfg: List[Tuple['SymbolData', DetectionConfig]], words: List[TextWord]) -> None:
        spatial_index = SpatialTextIndex(words)
        
        best_for_word_func = {} 
        best_for_word_mark = {}
        best_for_word_loc = {}

        for sym_idx, (sym, cfg) in enumerate(symbols_with_cfg):
            cx, cy = sym.center
            radius = cfg.search_radius
            nearby_words = spatial_index.query_nearby(sym.center, radius)

            for w in nearby_words:
                w_idx = w.original_index
                wx, wy = w.center_scaled
                print(w.text)

                dist = math.hypot(cx - wx, cy - wy)
                if dist > radius: continue
                
                if cfg._re_function and cfg._re_function.match(w.text):
                    prev = best_for_word_func.get(w_idx)
                    if prev is None or dist < prev[1]:
                        best_for_word_func[w_idx] = (sym_idx, dist)
                
                if cfg._re_mark and cfg._re_mark.match(w.text):
                    prev = best_for_word_mark.get(w_idx)
                    print(w.text)
                    if prev is None or dist < prev[1]:
                        best_for_word_mark[w_idx] = (sym_idx, dist)

                if cfg._re_location and cfg._re_location.match(w.text):
                    prev = best_for_word_loc.get(w_idx)
                    if prev is None or dist < prev[1]:
                        best_for_word_loc[w_idx] = (sym_idx, dist)

        def apply_map(mapping, attr_name, compiled_attr_name):
            for w_idx, (sym_idx, dist) in mapping.items():
                sym, cfg = symbols_with_cfg[sym_idx]
                pattern = getattr(cfg, compiled_attr_name)
                m = pattern.match(words[w_idx].text)
                value = m.group(1) if m and m.groups() else words[w_idx].text
                setattr(sym, attr_name, value)

        apply_map(best_for_word_func, "function", "_re_function")
        apply_map(best_for_word_mark, "mark", "_re_mark")
        apply_map(best_for_word_loc, "location", "_re_location")

    @staticmethod
    def assign_unique_terminals(symbols_with_cfg: List[Tuple['SymbolData', DetectionConfig]],
                                words: List[TextWord]) -> None:
        groups: Dict[str, List[int]] = {"terminal": [], "connector": [], "other": []}

        for idx, (sym, cfg) in enumerate(symbols_with_cfg):
            name = sym.export_macro.lower()
            if "terminal" in name:
                groups["terminal"].append(idx)
            elif "conn" in name:
                groups["connector"].append(idx)
            else:
                groups["other"].append(idx)

        for group_name in ("connector", "terminal", "other"):
            idxs = groups[group_name]
            if not idxs:
                continue
            TextService._assign_unique_terminals_for_indices(symbols_with_cfg, words, idxs)

    @staticmethod
    def _assign_unique_terminals_for_indices(symbols_with_cfg: List[Tuple['SymbolData', DetectionConfig]],
                                             words: List[TextWord],
                                             indices: List[int]) -> None:
        spatial_index = SpatialTextIndex(words)
        candidates = []

        for sym_idx in indices:
            sym, cfg = symbols_with_cfg[sym_idx]
            cx, cy = sym.center
            radius = cfg.search_radius
            nearby_words = spatial_index.query_nearby(sym.center, radius)

            for w in nearby_words:
                w_idx = w.original_index
                wx_raw, wy_raw = w.center_scaled
                dist = math.hypot(cx - wx_raw, cy - wy_raw)
                if dist > radius:
                    continue

                if cfg._re_term_num and cfg._re_term_num.match(w.text):
                    candidates.append({
                        'dist': dist, 'w_idx': w_idx, 's_idx': sym_idx,
                        'text': w.text, 'type': 'terminal',
                        'coord': (int(wx_raw), int(wy_raw)),
                        'angle': w.angle
                    })
                elif cfg._re_term_child and cfg._re_term_child.match(w.text):
                    candidates.append({
                        'dist': dist, 'w_idx': w_idx, 's_idx': sym_idx,
                        'text': w.text, 'type': 'parent',
                        'coord': (int(wx_raw), int(wy_raw)),
                        'angle': w.angle
                    })

        candidates.sort(key=lambda x: x['dist'])
        used_word_indices = set()
        sym_terminals = {i: [] for i in indices}

        for c in candidates:
            if c['w_idx'] in used_word_indices:
                continue
            used_word_indices.add(c['w_idx'])
            sym_idx = c['s_idx']
            sym_data, cfg = symbols_with_cfg[sym_idx]

            if c['type'] == 'parent':
                if sym_data.terminal_parent == "\u00A0":
                    sym_data.terminal_parent = c['text']
                    sym_data.terminal_parent_point = c['coord']
                    sym_data.terminal_parent_angle = int(c['angle'])
            else:
                wx, wy = c['coord']
                calc_x = wx - 5 if cfg.num_terminals == 4 else wx

                cx, cy = sym_data.center
                angle = math.degrees(math.atan2(-(wy - cy), calc_x - cx)) % 360
                sym_terminals.setdefault(sym_idx, []).append(
                    (c['text'], angle, c['dist'], c['coord'], c['angle'])
                )

        for sym_idx, terminals in sym_terminals.items():
            sym_data, cfg = symbols_with_cfg[sym_idx]
            if not terminals:
                continue

            terminals.sort(key=lambda x: x[2])
            if len(terminals) > cfg.num_terminals:
                terminals = terminals[:cfg.num_terminals]

            targets = {}
            rot = sym_data.rotation
            if cfg.num_terminals == 1:
                t = terminals[0]
                sym_data.terminals_found[t[0]] = "up"
                sym_data.terminals_exact_coords[t[0]] = (t[3][0], t[3][1], int(t[4]))
                continue
            elif cfg.num_terminals == 2:
                if rot in [90, 270]:
                    targets = {180: 'left', 0: 'right'}
                else:
                    targets = {90: 'up', 270: 'down'}
            elif cfg.num_terminals == 4:
                targets = {0: 'right', 90: 'up', 180: 'left', 270: 'down'}

            matches = []
            options_count = {t[0]: 0 for t in terminals}

            for t_text, t_ang, _, t_coord, t_text_angle in terminals:
                for slot_ang, slot_label in targets.items():
                    diff = min(abs(t_ang - slot_ang), 360 - abs(t_ang - slot_ang))
                    if diff <= 45:
                        matches.append({
                            'diff': diff, 'text': t_text, 'slot': slot_label,
                            'coord': t_coord, 'angle': t_text_angle
                        })
                        options_count[t_text] += 1

            matches.sort(key=lambda x: (options_count[x['text']], x['diff']))

            used_terminals = set()
            filled_slots = set()

            for m in matches:
                if m['text'] in used_terminals or m['slot'] in filled_slots:
                    continue
                sym_data.terminals_found[m['text']] = m['slot']
                sym_data.terminals_exact_coords[m['text']] = (
                    m['coord'][0], m['coord'][1], int(m['angle'])
                )
                used_terminals.add(m['text'])
                filled_slots.add(m['slot'])

            remaining_terms = [t for t in terminals if t[0] not in used_terminals]
            remaining_slots = {ang: lbl for ang, lbl in targets.items() if lbl not in filled_slots}

            if remaining_terms and remaining_slots:
                force_matches = []
                for t_text, t_ang, _, t_coord, t_text_angle in remaining_terms:
                    for slot_ang, slot_label in remaining_slots.items():
                        diff = min(abs(t_ang - slot_ang), 360 - abs(t_ang - slot_ang))
                        force_matches.append({
                            'diff': diff, 'text': t_text, 'slot': slot_label,
                            'coord': t_coord, 'angle': t_text_angle
                        })

                force_matches.sort(key=lambda x: x['diff'])

                for m in force_matches:
                    if m['text'] in used_terminals or m['slot'] in filled_slots:
                        continue
                    sym_data.terminals_found[m['text']] = m['slot']
                    sym_data.terminals_exact_coords[m['text']] = (
                        m['coord'][0], m['coord'][1], int(m['angle'])
                    )
                    used_terminals.add(m['text'])
                    filled_slots.add(m['slot'])

class HeuristicsService:
    @staticmethod
    def apply_all(symbols: List[SymbolData]) -> List[SymbolData]:
        valid_symbols = []
        for sym in symbols:
            if not HeuristicsService._check_terminal_validity(sym):
                continue
            valid_symbols.append(sym)
        
        valid_symbols = HeuristicsService._apply_implied_assignments(valid_symbols)
        return valid_symbols

    @staticmethod
    def _check_terminal_validity(sym: SymbolData) -> bool:
        return True

    @staticmethod
    def _apply_implied_assignments(symbols: List[SymbolData]) -> List[SymbolData]:
        pages = defaultdict(list)
        for s in symbols:
            pages[s.page_num].append(s)

        processed_symbols = []

        for page_num, page_syms in pages.items():
            terminals = [s for s in page_syms if "terminal" in s.export_macro.lower()]
            connectors = [s for s in page_syms if "conn" in s.export_macro.lower()]
            switches = [s for s in page_syms if "switch" in s.export_macro.lower()]
            others = [s for s in page_syms if s not in terminals and s not in connectors]

            HeuristicsService._assign_with_mode_locking(terminals)
            HeuristicsService._assign_with_mode_locking(switches)
            HeuristicsService._assign_with_growing_rectangles(connectors)

            processed_symbols.extend(terminals)
            processed_symbols.extend(connectors)
            processed_symbols.extend(others)

        return processed_symbols

    @staticmethod
    def _assign_with_growing_rectangles(group: List[SymbolData]):
        if not group: return

        headers = [s for s in group if s.mark and s.mark.strip()]
        targets = [s for s in group if not (s.mark and s.mark.strip())]
        if not headers or not targets: return

        targets.sort(key=lambda s: (s.center[1], s.center[0]))
        
        ALIGN_TOL = 15
        ORTHO_RANGE = 100

        def get_axis_info(sym):
            rot = int(sym.rotation % 360)
            if rot == 90 or rot == 270:
                return "vert", "horiz" 
            return "horiz", "vert"

        limits = {id(h): float('inf') for h in headers}

        for h in headers:
            p_axis, _ = get_axis_info(h)
            hx, hy = h.center
            
            for other in headers:
                if other is h: continue
                ox, oy = other.center
                
                if p_axis == "horiz":
                    if ox > hx and abs(oy - hy) < ALIGN_TOL:
                        limits[id(h)] = min(limits[id(h)], ox)
                else:
                    if oy > hy and abs(ox - hx) < ALIGN_TOL:
                        limits[id(h)] = min(limits[id(h)], oy)

        strip_members = []
        
        for t in targets:
            if t.assigned_header: continue 
            
            tx, ty = t.center
            best_header = None
            min_dist = float('inf')

            for h in headers:
                p_axis, _ = get_axis_info(h)
                hx, hy = h.center
                limit = limits[id(h)]
                
                in_rect = False
                
                if p_axis == "horiz":
                    if (hx - 5 < tx < limit) and (abs(ty - hy) < ALIGN_TOL):
                        in_rect = True
                else:
                    if (hy - 5 < ty < limit) and (abs(tx - hx) < ALIGN_TOL):
                        in_rect = True
                
                if in_rect:
                    dist = abs(hx - tx) + abs(hy - ty)
                    if dist < min_dist:
                        min_dist = dist
                        best_header = h
            
            if best_header:
                t.mark = best_header.mark
                t.is_inferred = True
                if best_header.location and best_header.location != "\u00A0":
                    t.location = best_header.location
                t.assigned_header = best_header
                strip_members.append(t)

        sources = headers + strip_members
        
        for parent in sources:
            px, py = parent.center
            _, ortho_axis = get_axis_info(parent)
            root = parent.assigned_header if parent.is_inferred else parent
            if not root: continue

            for t in targets:
                if t.assigned_header: continue
                
                tx, ty = t.center
                caught = False
                
                if ortho_axis == "vert":
                    if abs(tx - px) < ALIGN_TOL:
                        if abs(ty - py) < ORTHO_RANGE:
                            caught = True
                else:
                    if abs(ty - py) < ALIGN_TOL:
                        if abs(tx - px) < ORTHO_RANGE:
                            caught = True
                
                if caught:
                    t.mark = root.mark
                    t.is_inferred = True
                    if root.location and root.location != "\u00A0":
                        t.location = root.location
                    t.assigned_header = root

    @staticmethod
    def _assign_with_mode_locking(group: List[SymbolData]):
        if not group: return

        headers = [s for s in group if s.mark and s.mark.strip()]
        targets = [s for s in group if not (s.mark and s.mark.strip())]
        if not headers or not targets: return

        headers.sort(key=lambda s: (s.center[1], s.center[0]))
        assigned_ids = set()
        all_syms = headers + targets
        
        TOL = 25
        SEARCH_RADIUS = 300

        def is_clear(a, b, mode):
            ax, ay = a.center; bx, by = b.center
            for o in all_syms:
                if o is a or o is b: continue
                ox, oy = o.center
                if mode == "vert":
                    if min(ay, by) < oy < max(ay, by) and abs(ox - bx) < TOL: return False
                elif mode == "horiz":
                    if min(ax, bx) < ox < max(ax, bx) and abs(oy - by) < TOL: return False
            return True

        for h in headers:
            hx, hy = h.center
            
            best_t = None
            min_dist = float('inf')
            mode = None

            for t in targets:
                if id(t) in assigned_ids: continue
                tx, ty = t.center

                if tx < hx - 5 or ty < hy - 20: continue

                dx, dy = abs(hx - tx), abs(hy - ty)
                is_vert = dx < TOL
                is_horiz = dy < TOL
                if not is_vert and not is_horiz: continue

                dist = dx + dy
                if dist > SEARCH_RADIUS: continue

                curr_mode = "vert" if is_vert else "horiz"
                if not is_clear(h, t, curr_mode): continue

                if dist < min_dist:
                    min_dist = dist
                    best_t = t
                    mode = curr_mode

            if not best_t: continue

            queue = [best_t]
            while queue:
                curr = queue.pop(0)
                
                curr.mark = h.mark
                curr.is_inferred = True
                curr.assigned_header = h
                if h.location and h.location != "\u00A0": 
                    curr.location = h.location
                
                assigned_ids.add(id(curr))

                cx, cy = curr.center
                next_t = None
                next_dist = float('inf')

                for t in targets:
                    if id(t) in assigned_ids: continue
                    tx, ty = t.center

                    if tx < cx - 5 or ty < cy - 20: continue

                    dx, dy = abs(cx - tx), abs(cy - ty)
                    if mode == "vert" and dx > TOL: continue
                    if mode == "horiz" and dy > TOL: continue

                    dist = dx + dy
                    if dist > SEARCH_RADIUS: continue

                    if not is_clear(curr, t, mode): continue

                    if dist < next_dist:
                        next_dist = dist
                        next_t = t
                
                if next_t: queue.append(next_t)


# ─────────────────────────────────────────────────────────────────────────────
# 7. Processor
# ─────────────────────────────────────────────────────────────────────────────
class SymbolProcessor:
    def __init__(self, zoom: float = 3.0):
        self.zoom = zoom

    def process_page(self, pdf_path: str, page_num: int, configs: List[DetectionConfig]) -> List[SymbolData]:
        try:
            img, words, _ = PDFService.render_page(pdf_path, page_num, self.zoom)
        except Exception as e:
            print(f"Error reading page {page_num}: {e}")
            return []

        bbox_pdf = (0, 0, img.shape[1], img.shape[0])
        x0_px = int(bbox_pdf[0] * self.zoom)
        y0_px = int(bbox_pdf[1] * self.zoom)
        x1_px = int(bbox_pdf[2] * self.zoom)
        y1_px = int(bbox_pdf[3] * self.zoom)
        
        img_h, img_w = img.shape[:2]
        x0_px, y0_px = max(0, x0_px), max(0, y0_px)
        x1_px, y1_px = min(img_w, x1_px), min(img_h, y1_px)
        
        scene_crop = img[y0_px:y1_px, x0_px:x1_px]
        scene_crop_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        symbols_with_cfg = []
        for cfg in configs:
            raw_matches = VisionService.match_precomputed(
                scene_crop_gray, cfg._precomputed_variants, cfg.threshold
            )
            for m in raw_matches:
                sym = SymbolData(
                    template_name=Path(cfg.template_path).stem,
                    export_macro=cfg.export_macro,
                    center=(m["center"][0] + x0_px, m["center"][1] + y0_px),
                    bbox=(m["bbox"][0] + x0_px, m["bbox"][1] + y0_px, m["bbox"][2] + x0_px, m["bbox"][3] + y0_px),
                    score=m["score"],
                    rotation=m["rotation"],
                    page_num=page_num + 1,
                    flipped=m["flipped"],
                )
                symbols_with_cfg.append((sym, cfg))

        symbols_with_cfg = VisionService.cross_class_nms(
            symbols_with_cfg, 
            iou_threshold=0.3
        )

        TextService.assign_unique_labels(symbols_with_cfg, words)
        TextService.assign_unique_terminals(symbols_with_cfg, words)

        results = []
        for sym, cfg in symbols_with_cfg:
            if cfg.label_required:
                if cfg._re_function and (not sym.function or not sym.function.strip() or sym.function == "\u00A0"): continue
                if cfg._re_mark and (not sym.mark or not sym.mark.strip()): continue
            results.append(sym)

        results = HeuristicsService.apply_all(results)

        return results

# ─────────────────────────────────────────────────────────────────────────────
# 8. Debug Visualizer
# ─────────────────────────────────────────────────────────────────────────────
class DebugVisualizer:
    @staticmethod
    def show_window(pdf_path: str, page_num: int, zoom: float,
                    symbols: List[SymbolData], words: List[TextWord]):
        print(f"Visualizing Page {page_num}...")

        image, _, _ = PDFService.render_page(pdf_path, page_num - 1, zoom)

        vis_img = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        h, w = vis_img.shape[:2]

        fig, ax = plt.subplots(figsize=(min(14, w / 70), min(10, h / 70)))
        ax.imshow(vis_img)
        ax.set_title(f"Page {page_num}", fontsize=14)
        ax.axis("off")

        for wobj in words:
            x0, y0, x1, y1 = wobj.rect
            rect = Rectangle(
                (x0, y0), x1 - x0, y1 - y0,
                linewidth=0.4, edgecolor="lightgray", facecolor="none", alpha=0.0,
            )
            ax.add_patch(rect)

        family_map: Dict[int, List[SymbolData]] = defaultdict(list)
        for s in symbols:
            if s.assigned_header:
                family_map[id(s.assigned_header)].append(s)
            elif s.mark and s.location:
                family_map[id(s)].append(s)

        unique_templates = list(set(s.template_name for s in symbols))
        colors = {}
        for temp in unique_templates:
            seed = (hash(temp) & 0xFFFFFFFF)
            rng = random.Random(seed)
            colors[temp] = (rng.random(), rng.random(), rng.random())

        for header_id, members in family_map.items():
            if not members: continue
            min_x = min(s.bbox[0] for s in members)
            min_y = min(s.bbox[1] for s in members)
            max_x = max(s.bbox[2] for s in members)
            max_y = max(s.bbox[3] for s in members)

            fam_rect = Rectangle(
                (min_x - 10, min_y - 10),
                (max_x - min_x) + 20, (max_y - min_y) + 20,
                linewidth=2.0, edgecolor="magenta", facecolor="none", linestyle="--", alpha=0.9,
            )
            ax.add_patch(fam_rect)

            header = next((m for m in members if id(m) == header_id), None)
            if header:
                label = f"{header.mark}"
                ax.text(min_x, min_y - 15, label, fontsize=5, color="magenta",
                        bbox=dict(facecolor="white", edgecolor="magenta", linewidth=0.8, alpha=0.0))

        for sym in symbols:
            x1, y1, x2, y2 = sym.bbox
            cx, cy = sym.center
            color = colors.get(sym.template_name, (1.0, 0.0, 0.0))

            box_rect = Rectangle(
                (x1, y1), x2 - x1, y2 - y1,
                linewidth=2.0, edgecolor=color, facecolor="none", alpha=0.0,
            )
            ax.add_patch(box_rect)

            ax.plot(cx, cy, marker="o", markersize=6, markeredgecolor="black",
                    markerfacecolor="yellow", alpha=0.0)

            if sym.assigned_header:
                hx, hy = sym.assigned_header.center
                arrow = FancyArrowPatch(
                    (hx, hy), (cx, cy),
                    arrowstyle="->", mutation_scale=10, linewidth=1.8,
                    edgecolor="lime", facecolor="lime", alpha=0.85,
                )
                ax.add_patch(arrow)

            label = f"{sym.template_name} ({sym.score:.2f})"
            ax.text(cx, cy - 10, label, fontsize=5, color="black", ha="center", va="bottom",
                    bbox=dict(facecolor="white", edgecolor=color, linewidth=0.8, alpha=0.8))

            attr_lines = []
            if sym.function and sym.function != "\u00A0":
                attr_lines.append(f"Fn: {sym.function}")
            if sym.mark:
                attr_lines.append(f"Mk: {sym.mark}")
            if sym.location:
                attr_lines.append(f"Loc: {sym.location}")
            if sym.terminal_parent and sym.terminal_parent.strip() and sym.terminal_parent != "\u00A0":
                attr_lines.append(f"P: {sym.terminal_parent}")
            if sym.is_inferred:
                attr_lines.append("(Inferred)")

            text_y = y2 + 15
            for line in attr_lines:
                ax.text(x1, text_y, line, fontsize=5, color="black",
                        bbox=dict(facecolor="black", edgecolor="none", alpha=0.0))
                text_y += 12

            offsets = {"up": (0, -40), "down": (0, 40), "left": (-40, 0), "right": (40, 0)}
            
            for term_text, direction in sym.terminals_found.items():
                off_x, off_y = offsets.get(direction, (0, 0))
                ex, ey = cx + off_x, cy + off_y

                ax.plot([cx, ex], [cy, ey], linewidth=1.5, color="cyan", alpha=0.85)
                ax.text(ex, ey, term_text, fontsize=5, color="black", ha="center", va="center",
                        bbox=dict(facecolor="white", edgecolor="cyan", linewidth=0.8, alpha=0.9))

        plt.tight_layout()
        plt.show()
        return None

# ─────────────────────────────────────────────────────────────────────────────
# 9. Excel Writer & Main
# ─────────────────────────────────────────────────────────────────────────────
class ExcelExporter:
    def __init__(self, template_path: str, output_path: str):
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)

    def write(self, symbols: List[SymbolData], page_heights: Dict[int, float], zoom: float):
        if not self.template_path.exists():
            wb = Workbook(); ws = wb.active
        else:
            wb = load_workbook(self.template_path); ws = wb.active 

        row = ws.max_row + 1
        while row > 1 and ws.cell(row=row-1, column=1).value is None: row -= 1
        
        # ─────────────────────────────────────────────────────────────────
        # COORDINATE SYSTEM: Inches (matching wires/DXF coordinate system)
        #
        # PDF uses 72 points per inch. All pixel coordinates are converted
        # to inches by: inches = pixel / zoom / 72
        #
        # Y is flipped so Y=0 is at the bottom of the page (matching
        # CAD / macro coordinate conventions used by the wires script).
        # ─────────────────────────────────────────────────────────────────
        
        PTS_PER_INCH = 72.0
        
        print(f"Writing data... processing {len(symbols)} symbols...")
        print(f"Coordinate system: inches, zoom={zoom}")
        
        def fmt(txt): 
            if not txt or not isinstance(txt, str): return "\u00A0"
            txt = txt.strip()
            if not txt: return "\u00A0"
            return txt.replace("=", " =").replace("+", " +")

        for s in symbols:
            
            # Main Symbol Row
            clean_macro = re.sub(r"\s+", "_", s.export_macro).lower()
            if s.flipped:
                clean_macro = f"{clean_macro}_flipped"
            if s.is_inferred:
                clean_macro = f"{clean_macro}_implied"
            symbol_macro_name = f"{clean_macro}_{int(s.rotation)}"
            
            # page_h is in PDF points (from page.rect.height)
            page_h_pt = page_heights.get(s.page_num - 1, 842)
            page_h_in = page_h_pt / PTS_PER_INCH
            
            # Convert pixel coords → inches
            sym_x_in = s.center[0] / zoom / PTS_PER_INCH
            sym_y_in = page_h_in - (s.center[1] / zoom / PTS_PER_INCH)

            ws.cell(row=row, column=1, value=symbol_macro_name)
            ws.cell(row=row, column=2, value=round(sym_x_in, 4))
            ws.cell(row=row, column=3, value=round(sym_y_in, 4))
            ws.cell(row=row, column=6, value=s.page_num)
            ws.cell(row=row, column=11, value=fmt(s.location))
            
            raw_function = s.function
            if raw_function and raw_function.strip().startswith("="):
                raw_function = raw_function.split('-')[0].strip()
            
            ws.cell(row=row, column=13, value=fmt(raw_function))
            ws.cell(row=row, column=23, value=fmt(s.mark))
            ws.cell(row=row, column=24, value="\u00A0") 
            ws.cell(row=row, column=25, value="\u00A0") 
            ws.cell(row=row, column=26, value="\u00A0") 
            ws.cell(row=row, column=27, value="\u00A0") 
            ws.cell(row=row, column=28, value="\u00A0") 
            row += 1

            # Child Terminals
            for term_text, (tx_px, ty_px, raw_angle) in s.terminals_exact_coords.items():
                raw_angle = raw_angle % 360
                rotation = int(round(raw_angle / 90) * 90) % 360
                if rotation == 270: suffix = "90"
                elif rotation == 90: suffix = "270"
                else: suffix = str(rotation)
                
                macro_name = f"text_{suffix}"

                # Convert pixel coords → inches
                term_x_in = tx_px / zoom / PTS_PER_INCH
                term_y_in = page_h_in - (ty_px / zoom / PTS_PER_INCH)
                
                ws.cell(row=row, column=1, value=macro_name)
                ws.cell(row=row, column=2, value=round(term_x_in, 4))
                ws.cell(row=row, column=3, value=round(term_y_in, 4))
                ws.cell(row=row, column=6, value=s.page_num)
                ws.cell(row=row, column=11, value="\u00A0")
                ws.cell(row=row, column=13, value="\u00A0")
                ws.cell(row=row, column=23, value="\u00A0")
                ws.cell(row=row, column=24, value=fmt(term_text))
                row += 1

            # Parent Terminal
            if s.terminal_parent and s.terminal_parent.strip() and s.terminal_parent != "\u00A0" and s.terminal_parent_point:
                px_px, py_px = s.terminal_parent_point
                raw_angle = s.terminal_parent_angle
                raw_angle = raw_angle % 360
                rotation = int(round(raw_angle / 90) * 90) % 360
                if rotation == 270: suffix = "90"
                elif rotation == 90: suffix = "270"
                else: suffix = str(rotation)

                macro_name = f"text_{suffix}"

                # Convert pixel coords → inches
                parent_x_in = px_px / zoom / PTS_PER_INCH
                parent_y_in = page_h_in - (py_px / zoom / PTS_PER_INCH)

                ws.cell(row=row, column=1, value=macro_name)
                ws.cell(row=row, column=2, value=round(parent_x_in, 4))
                ws.cell(row=row, column=3, value=round(parent_y_in, 4))
                ws.cell(row=row, column=6, value=s.page_num)
                ws.cell(row=row, column=11, value="\u00A0")
                ws.cell(row=row, column=13, value="\u00A0")
                ws.cell(row=row, column=23, value="\u00A0")
                ws.cell(row=row, column=24, value=fmt(s.terminal_parent))
                row += 1

        wb.save(self.output_path)
        print(f"Saved to {self.output_path} (coordinates in inches)")

def worker_process_page(args):
    pdf, p_num, zoom, cfgs = args
    proc = SymbolProcessor(zoom)
    _, words, dims = PDFService.render_page(pdf, p_num, zoom)
    syms = proc.process_page(pdf, p_num, cfgs)
    return p_num, dims[1], syms, words

def main():
    # --- CONFIGURATION ---
    PDF_FILE = "PDF\C2 LV Schematics allocr.pdf"
    XLSX_TEMPLATE = "C:/Users/John.Hoang/Documents/lol_out/template.xlsx"
    XLSX_OUTPUT = "C:/Users/John.Hoang/Documents/lol_out/output_final.xlsx"
    ZOOM_LEVEL = 3.0 
    DEBUG_MODE = True
    configurations = [

        # DetectionConfig(
        #     template_path="symbols/no_power_switch.png",
        #     export_macro="no_power_switch",
        #     threshold=0.6,
        #     search_radius=350,
        #     overlap_threshold=0.5,
        #     rotate_vary=False,
        #     flip_vary=False,
        #     num_terminals=2,
        #     label_required=False,
        #     function_regex = None,
        #     mark_regex=r"^(\d{2}[A-Za-z]\d+).*$",
        #     location_regex=r"^[+ 4]{1,2}([A-Z]{3}).*$",
        #     terminal_number_regex = r"^([a-zA-Z]?\d{1,2}[a-zA-Z]?|[a-zA-Z]\d{1,2}[+-]?)$",
        #     use_mask=False
        # ),

        DetectionConfig(
            template_path="symbols/fuse.png",
            export_macro="fuse",
            threshold=0.7,
            search_radius=250,
            overlap_threshold=0.2,
            rotate_vary=True,
            num_terminals=2,
            label_required=False,
            function_regex=None,
            mark_regex=r"^(\d{2}[F]\d+).*$",
            location_regex=r"^[+4]\s*([A-Z]{3}).*$",
            terminal_number_regex=r"^([a-zA-Z]?\d{1,2}[a-zA-Z]?|[a-zA-Z]\d{1,2}[+-]?)$",
            terminal_child_regex=None,
            use_mask= False,
        ),
    ]

    print("Pre-computing template rotations...")
    for cfg in configurations:
        cfg.precompute_templates()

    pages_to_scan = list(range(22,492)) 
    
    all_symbols = []
    page_heights_map = {}
    
    t0 = time.time()

    if DEBUG_MODE:
        pg = 1
        pages_to_scan = list(range(pg-1,pg))
        print(f"--- DEBUG MODE ON ---")
        print("Running sequentially. Close window to continue.")
        for p in pages_to_scan:
            args = (PDF_FILE, p, ZOOM_LEVEL, configurations)
            p_num, height, syms, words = worker_process_page(args)
            all_symbols.extend(syms)
            page_heights_map[p_num] = height
            print(f"Page {p_num+1}: Found {len(syms)} symbols.")
            key = DebugVisualizer.show_window(PDF_FILE, p_num+1, ZOOM_LEVEL, syms, words)
            if key == 27: break
    else:
        print(f"--- DEBUG MODE OFF ---")
        max_workers = min(os.cpu_count() or 1, 6)
        chunk_size = max(1, len(pages_to_scan) // (max_workers * 4))
        tasks = [(PDF_FILE, p, ZOOM_LEVEL, configurations) for p in pages_to_scan]
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            for p_num, height, syms, _ in executor.map(worker_process_page, tasks, chunksize=chunk_size):
                if len(syms) > 0:
                    print(f"Page {p_num+1}: Found {len(syms)} symbols.")
                all_symbols.extend(syms)
                page_heights_map[p_num] = height

    exporter = ExcelExporter(XLSX_TEMPLATE, XLSX_OUTPUT)
    exporter.write(all_symbols, page_heights_map, ZOOM_LEVEL)
    print(f"Total time: {time.time() - t0:.2f}s")

if __name__ == "__main__":
    main()