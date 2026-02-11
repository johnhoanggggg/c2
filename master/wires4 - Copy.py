import io
import json
import math
import re
import time
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple, Set, Dict
from concurrent.futures import ProcessPoolExecutor
import multiprocessing

import ezdxf
import numpy as np
from PIL import Image
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.patches as patches

# Try to import pytesseract
try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False
    print("Warning: pytesseract not installed. OCR will be skipped.")


# -----------------------
# Domain Model
# -----------------------

@dataclass
class Wire:
    """Represents a detected wire with its associated metadata."""
    segments: List[Tuple[Tuple[float, float], Tuple[float, float]]] = field(default_factory=list)
    label: Optional[str] = None
    equip: Optional[str] = None
    mark: Optional[str] = None
    wire_spec: Optional[str] = None
    wirematch_spec: Optional[str] = None

    spec_source: Optional[str] = None

    mark_pos_raw: Optional[Tuple[float, float]] = None
    mark_pos: Optional[Tuple[float, float]] = None
    mark_orientation: Optional[str] = None
    mark_angle_deg: Optional[float] = None

    _bbox: Optional[Tuple[float, float, float, float]] = None

    def add_segment(self, seg: Tuple[Tuple[float, float], Tuple[float, float]]) -> None:
        self.segments.append(seg)
        self._bbox = None

    def get_bbox(self) -> Tuple[float, float, float, float]:
        if self._bbox:
            return self._bbox
        if not self.segments:
            return (0, 0, 0, 0)
        xs = [p[0] for seg in self.segments for p in seg]
        ys = [p[1] for seg in self.segments for p in seg]
        self._bbox = (min(xs), min(ys), max(xs), max(ys))
        return self._bbox

    def set_label(self, label: Optional[str]) -> None:
        self.label = label
        if label and '.' in label:
            parts = label.split('.')
            if len(parts) == 2:
                self.equip, self.mark = parts[0], parts[1]
            else:
                self.equip, self.mark = parts[0], parts[-1]
        else:
            self.equip, self.mark = label, None


# -----------------------
# Utility Functions
# -----------------------

def get_nearest_point_on_segment(px, py, seg_start, seg_end):
    x0, y0 = seg_start
    x1, y1 = seg_end
    vx, vy = x1 - x0, y1 - y0
    wx, wy = px - x0, py - y0
    det = vx * vx + vy * vy
    if det == 0:
        return x0, y0
    c1 = vx * wx + vy * wy
    if c1 <= 0:
        return x0, y0
    if det <= c1:
        return x1, y1
    t = c1 / det
    return x0 + t * vx, y0 + t * vy


def pt_seg_dist(px, py, seg_start, seg_end):
    nx, ny = get_nearest_point_on_segment(px, py, seg_start, seg_end)
    return math.hypot(px - nx, py - ny)


def get_segment_orientation(seg):
    (x0, y0), (x1, y1) = seg
    if abs(y1 - y0) < 1e-3:
        return "H", ((min(x0, x1), y0), (max(x0, x1), y1)), abs(x1 - x0)
    elif abs(x1 - x0) < 1e-3:
        return "V", ((x0, min(y0, y1)), (x1, max(y0, y1))), abs(y1 - y0)
    return None, seg, 0


def generate_chunk_starts(start_pt, kind, length, seg_unit):
    sx, sy = start_pt
    dx = seg_unit if kind == "H" else 0
    dy = seg_unit if kind == "V" else 0
    curx, cury = sx, sy
    while (kind == "H" and curx + seg_unit < sx + length) or \
          (kind == "V" and cury + seg_unit < sy + length):
        yield curx, cury
        curx += dx
        cury += dy
    if length % seg_unit:
        yield (sx + length - seg_unit, sy) if kind == "H" else (sx, sy + length - seg_unit)


def get_rotation_vector(angle_deg):
    rad = math.radians(angle_deg)
    return -math.sin(rad), -math.cos(rad)


def is_vertical_text(angle_deg):
    a = abs(angle_deg) % 180
    return 45 < a < 135


# -----------------------
# OCR Helper
# -----------------------

class DXFOcrHelper:
    """Renders a DXF to an image and performs OCR to extract text with positions in DXF coords."""

    def __init__(self, dpi: int = 300):
        self.dpi = dpi

    def render_dxf_to_image(self, doc):
        """Render the DXF modelspace to a PIL image. Returns (image, transform_info)."""
        from ezdxf.addons.drawing import RenderContext, Frontend
        from ezdxf.addons.drawing.matplotlib import MatplotlibBackend

        fig = plt.figure(figsize=(11, 8.5), dpi=self.dpi)
        ax = fig.add_axes([0, 0, 1, 1])
        ax.set_facecolor('white')

        ctx = RenderContext(doc)
        out = MatplotlibBackend(ax)
        Frontend(ctx, out).draw_layout(doc.modelspace(), finalize=True)

        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ax.axis('off')

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=self.dpi,
                    facecolor='white', edgecolor='none', pad_inches=0)
        plt.close(fig)

        buf.seek(0)
        img = Image.open(buf).convert('RGB')

        transform_info = {
            'xlim': xlim,
            'ylim': ylim,
            'img_width': img.width,
            'img_height': img.height,
            'dpi': self.dpi,
        }
        return img, transform_info

    def pixel_to_dxf_coords(self, px_x, px_y, transform_info):
        """Convert pixel coordinates to DXF coordinate space."""
        img_w = transform_info['img_width']
        img_h = transform_info['img_height']
        x_min, x_max = min(transform_info['xlim']), max(transform_info['xlim'])
        y_min, y_max = min(transform_info['ylim']), max(transform_info['ylim'])

        norm_x = px_x / img_w
        norm_y = px_y / img_h

        dxf_x = x_min + norm_x * (x_max - x_min)
        dxf_y = y_max - norm_y * (y_max - y_min)  # image Y is flipped vs DXF Y
        return dxf_x, dxf_y

    def perform_ocr(self, doc):
        """Run OCR at multiple rotations. Returns list of dicts with text, dxf_x, dxf_y, confidence, rotation."""
        if not TESSERACT_AVAILABLE:
            print("  OCR skipped (pytesseract not installed)")
            return []

        print("  Rendering DXF for OCR...")
        img, tinfo = self.render_dxf_to_image(doc)
        if img is None:
            return []

        # Tesseract config — allow digits + alpha + common label chars
        config = '--psm 11 -c tessedit_char_whitelist=0123456789'

        all_items = []
        rotations = [(0, "normal"), (90, "90° CCW"), (270, "90° CW")]

        for angle, desc in rotations:
            print(f"  OCR pass: {desc}...")
            rotated = img if angle == 0 else img.rotate(angle, expand=True)
            ocr_data = pytesseract.image_to_data(rotated, config=config,
                                                  output_type=pytesseract.Output.DICT)

            for i in range(len(ocr_data['text'])):
                text = ocr_data['text'][i].strip()
                if not text:
                    continue
                confidence = int(ocr_data['conf'][i])
                if confidence < 20:
                    continue

                cx = ocr_data['left'][i] + ocr_data['width'][i] / 2
                cy = ocr_data['top'][i] + ocr_data['height'][i] / 2

                # Map rotated pixel coords back to original image pixel coords
                orig_w, orig_h = img.width, img.height
                if angle == 0:
                    opx, opy = cx, cy
                elif angle == 90:
                    opx, opy = orig_w - cy, cx
                elif angle == 270:
                    opx, opy = cy, orig_h - cx
                else:
                    opx, opy = cx, cy

                dxf_x, dxf_y = self.pixel_to_dxf_coords(opx, opy, tinfo)

                all_items.append({
                    'text': text,
                    'dxf_x': dxf_x,
                    'dxf_y': dxf_y,
                    'confidence': confidence,
                    'rotation': angle,
                })

        # Deduplicate
        unique = []
        pos_thresh = 0.001  # DXF units (inches)
        for item in all_items:
            dup = False
            for ex in unique:
                if ex['text'] == item['text']:
                    if abs(ex['dxf_x'] - item['dxf_x']) < pos_thresh and \
                       abs(ex['dxf_y'] - item['dxf_y']) < pos_thresh:
                        if item['confidence'] > ex['confidence']:
                            unique.remove(ex)
                            unique.append(item)
                        dup = True
                        break
            if not dup:
                unique.append(item)

        print(f"  OCR found {len(unique)} unique text items")
        return unique


# -----------------------
# Main Processor Class
# -----------------------

class WireDetectionAndPlacement:
    MACROS = {
        0.25: ("hwire5", "vwire5"),
        0.5:  ("hwire10", "vwire10"),
        1.5:  ("hwire30", "vwire30"),
        2.5:  ("hwire50", "vwire50"),
    }

    def __init__(
        self,
        dxf_path: str,
        seg_min: float,
        connect_dist: float,
        label_dist: float,
        template_xlsx: str,
        file_name_value: 0,
        zoom: float = 3,
        wire_spec_json: Optional[str] = None,
        cable_spec_json: Optional[str] = None,
        ocr_dpi: int = 300,
    ) -> None:
        self.dxf_path = dxf_path
        self.seg_min = seg_min
        self.connect_dist = connect_dist
        self.label_dist = label_dist
        self.template_xlsx = Path(template_xlsx)
        self.file_name_value = file_name_value
        self.zoom = zoom  # only for debug visualization
        self.ocr_dpi = ocr_dpi

        self._wire_regex = re.compile(r'\d{5,}')
        self._cable_regex = re.compile(r'^-?[A-Za-z0-9]{7}$')
        self._wirematch_regex = re.compile(
            r'^[=+]?[A-Za-z0-9]{2}\s*\.\s*[A-Za-z0-9]{2}\s*\.\s*[A-Za-z0-9]{2}\s*\.\s*[A-Za-z0-9]{2}$')

        self.wire_spec_regex = self._load_and_compile_specs(wire_spec_json, match_type="exact")
        self.cable_spec_regex = self._load_and_compile_specs(cable_spec_json, match_type="contains")

        self.labels: List = []
        self.segments: List = []
        self.wires: List[Wire] = []
        self.spec_tags: List = []
        self.spec_tag_dist = label_dist
        self.wirematches: List = []
        self.matched_labels: List = []

    # ---- spec loading ----

    def _load_allowed_specs(self, path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return {s for s in json.load(f) if isinstance(s, str)}
        except (FileNotFoundError, json.JSONDecodeError):
            return set()

    def _load_and_compile_specs(self, json_path, match_type="exact"):
        if not json_path:
            return None
        specs = self._load_allowed_specs(json_path)
        if not specs:
            return None
        sorted_specs = sorted(specs, key=len, reverse=True)
        parts = [re.escape(s).replace(r"\ ", r"\s+") for s in sorted_specs]
        combined = '|'.join(parts)
        if match_type == "exact":
            return re.compile(f"^({combined})$")
        else:
            return re.compile(f"({combined})")

    def _reset_state(self):
        self.labels.clear()
        self.segments.clear()
        self.wires.clear()
        self.spec_tags.clear()
        self.matched_labels.clear()
        self.wirematches.clear()

    # ---- Extract segments from DXF directly ----

    def extract_segments_from_dxf(self, doc):
        """Read LINE entities from DXF modelspace. Coords are native DXF units (inches)."""
        self.segments = []
        msp = doc.modelspace()
        for entity in msp:
            if entity.dxftype() != "LINE":
                continue
            try:
                s = entity.dxf.start
                e = entity.dxf.end
                x0, y0 = s.x, s.y
                x1, y1 = e.x, e.y
            except Exception:
                continue

            length = math.hypot(x1 - x0, y1 - y0)
            if length < self.seg_min:
                continue

            # Only keep horizontal / vertical lines
            if abs(y1 - y0) <= 0.01 or abs(x1 - x0) <= 0.01:
                # Normalize direction
                if x0 > x1 or (x0 == x1 and y0 > y1):
                    self.segments.append(((x1, y1), (x0, y0)))
                else:
                    self.segments.append(((x0, y0), (x1, y1)))

        print(f"  Extracted {len(self.segments)} line segments from DXF")

    # ---- Extract labels from OCR results ----

    def extract_labels_from_ocr(self, ocr_items):
        """Classify OCR text items into wire labels, spec tags, and wirematches."""
        for item in ocr_items:
            word = item['text'].strip()
            cx, cy = item['dxf_x'], item['dxf_y']
            # Estimate angle from OCR rotation
            angle_deg = 0.0
            if item.get('rotation') == 90:
                angle_deg = 90.0
            elif item.get('rotation') == 270:
                angle_deg = -90.0

            is_label = False
            if self._wire_regex.match(word) or self._cable_regex.match(word):
                self.labels.append((cx, cy, word, angle_deg))
                is_label = True

            if not is_label:
                found_spec = False
                if self.cable_spec_regex and self.cable_spec_regex.search(word):
                    self.spec_tags.append((cx, cy, word, "cable"))
                    found_spec = True
                if not found_spec and self.wire_spec_regex and self.wire_spec_regex.search(word):
                    self.spec_tags.append((cx, cy, word, "wire"))

            if self._wirematch_regex.match(word):
                self.wirematches.append((cx, cy, word))

        print(f"  Labels: {len(self.labels)}, Spec tags: {len(self.spec_tags)}, "
              f"Wire matches: {len(self.wirematches)}")

    # ---- Segment grouping ----

    def _segments_touch(self, seg_a, seg_b, r):
        # 1. Quick Bounding Box check (Keep this for performance)
        ax_min, ax_max = min(seg_a[0][0], seg_a[1][0]), max(seg_a[0][0], seg_a[1][0])
        ay_min, ay_max = min(seg_a[0][1], seg_a[1][1]), max(seg_a[0][1], seg_a[1][1])
        bx_min, bx_max = min(seg_b[0][0], seg_b[1][0]), max(seg_b[0][0], seg_b[1][0])
        by_min, by_max = min(seg_b[0][1], seg_b[1][1]), max(seg_b[0][1], seg_b[1][1])
        
        if (ax_max + r < bx_min) or (bx_max + r < ax_min) or \
        (ay_max + r < by_min) or (by_max + r < ay_min):
            return False

        # 2. Check if any endpoint of A is near the line of B
        for p in seg_a:
            if pt_seg_dist(p[0], p[1], seg_b[0], seg_b[1]) <= r:
                return True
                
        # 3. Check if any endpoint of B is near the line of A
        for p in seg_b:
            if pt_seg_dist(p[0], p[1], seg_a[0], seg_a[1]) <= r:
                return True
                
        return False
    MACRO_LENGTHS = [2.5, 1.5, 0.5, 0.25]

    def tile_segment(self, kind, start_pt, length):
        """Greedy largest-first tiling. No overlap, no gaps."""
        length = math.floor(length / 0.05) * 0.05  # round down to nearest 0.05
        placements = []
        offset = 0.0

        while remaining := round(length - offset, 4):
            if remaining <= 0:
                break
            for unit in self.MACRO_LENGTHS:
                if unit <= remaining + 1e-6:
                    px = start_pt[0] + (offset if kind == "H" else 0)
                    py = start_pt[1] + (offset if kind == "V" else 0)
                    h_mac, v_mac = self.MACROS[unit]
                    placements.append((h_mac if kind == "H" else v_mac, px, py))
                    offset += unit
                    break

        return placements

    def group_segments(self):
        n = len(self.segments)
        if n == 0:
            self.wires = []
            return
        indexed = sorted(enumerate(self.segments), key=lambda x: min(x[1][0][0], x[1][1][0]))
        parent = list(range(n))

        def find(i):
            if parent[i] != i:
                parent[i] = find(parent[i])
            return parent[i]

        def unite(i, j):
            ri, rj = find(i), find(j)
            if ri != rj:
                parent[rj] = ri

        r = self.connect_dist
        for i in range(n):
            idx_i, seg_i = indexed[i]
            x_limit = max(seg_i[0][0], seg_i[1][0]) + r
            for j in range(i + 1, n):
                idx_j, seg_j = indexed[j]
                if min(seg_j[0][0], seg_j[1][0]) > x_limit:
                    break
                if self._segments_touch(seg_i, seg_j, r):
                    unite(idx_i, idx_j)

        clusters = {}
        for i in range(n):
            clusters.setdefault(find(i), []).append(i)

        self.wires = []
        for indices in clusters.values():
            wire = Wire()
            for idx in indices:
                wire.add_segment(self.segments[idx])
            self.wires.append(wire)

    # ---- Label matching ----

    def match_labels_to_wires(self):
        self.matched_labels.clear()
        wire_bboxes = [w.get_bbox() for w in self.wires]
        limit = self.label_dist

        for cx, cy, word, angle_deg in self.labels:
            nx, ny = get_rotation_vector(angle_deg)
            is_text_vertical = is_vertical_text(angle_deg)
            best_d = float("inf")
            best_wire = None
            for i, wire in enumerate(self.wires):
                wb = wire_bboxes[i]
                if cx < wb[0] - limit or cx > wb[2] + limit or \
                   cy < wb[1] - limit or cy > wb[3] + limit:
                    continue
                for seg in wire.segments:
                    kind, _, _ = get_segment_orientation(seg)
                    if kind is None:
                        continue
                    if is_text_vertical and kind == "H":
                        continue
                    if not is_text_vertical and kind == "V":
                        continue
                    nx_pt, ny_pt = get_nearest_point_on_segment(cx, cy, seg[0], seg[1])
                    ddx, ddy = nx_pt - cx, ny_pt - cy
                    if (ddx * nx + ddy * ny) <= 0:
                        continue
                    dist = math.hypot(cx - nx_pt, cy - ny_pt)
                    if dist < best_d:
                        best_d = dist
                        best_wire = wire

            if best_wire and best_d <= self.label_dist:
                best_wire.set_label(word)
                self.matched_labels.append((cx, cy, word, best_wire, angle_deg))
                best_wire.mark_pos_raw = (cx, cy)
                best_wire.mark_pos = (cx, cy)
                best_wire.mark_angle_deg = angle_deg
                best_wire.mark_orientation = "V" if is_text_vertical else "H"

        for cx, cy, word in self.wirematches:
            best_d, best_wire = float("inf"), None
            for i, wire in enumerate(self.wires):
                wb = wire_bboxes[i]
                if cx < wb[0] - limit or cx > wb[2] + limit or \
                   cy < wb[1] - limit or cy > wb[3] + limit:
                    continue
                for seg in wire.segments:
                    d = pt_seg_dist(cx, cy, seg[0], seg[1])
                    if d < best_d:
                        best_d, best_wire = d, wire
            if best_wire and best_d <= self.label_dist:
                best_wire.wirematch_spec = word

    def match_specs_within_radius_of_labels(self, radius=None, consume_each_spec=False):
        radius = radius or self.spec_tag_dist or self.label_dist
        if not self.spec_tags or not self.matched_labels:
            return
        remaining = list(self.spec_tags)
        for lx, ly, _, wire, angle_deg in self.matched_labels:
            candidates = []
            for idx, (sx, sy, stag, stype) in enumerate(remaining):
                if abs(sx - lx) > radius or abs(sy - ly) > radius:
                    continue
                dist = math.hypot(sx - lx, sy - ly)
                if dist > radius:
                    continue
                candidates.append((dist, idx, stag, stype))
            if candidates:
                candidates.sort(key=lambda t: t[0])
                _, best_idx, best_tag, best_source = candidates[0]
                wire.wire_spec = best_tag
                wire.spec_source = best_source
                if consume_each_spec:
                    remaining.pop(best_idx)

    def filter_wires(self):
        self.wires = [w for w in self.wires
                      if w.label or w.equip or w.mark or w.wirematch_spec]

    # ---- Build output rows ----

    def build_rows(self):
        rows = []
        for wire in self.wires:
            # 1. Wire Segment Macros
            for seg in wire.segments:
                kind, seg_ord, length = get_segment_orientation(seg)
                if not kind:
                    continue
                if length >= 2.5:
                    seg_unit = 2.5
                elif length >= 1.5:
                    seg_unit = 1.5
                elif length >= 0.5:
                    seg_unit = 0.5
                else:
                    seg_unit = 0.25

                h_mac, v_mac = self.MACROS[seg_unit]
                mac_name = h_mac if kind == "H" else v_mac

                sx, sy = seg_ord[0]
                for px, py in generate_chunk_starts((sx, sy), kind, length, seg_unit):
                    rows.append({
                        "#mac_name": mac_name,
                        "#mac_posx": f"{round(px, 2)}",
                        "#mac_posy": f"{round(py, 2)}",
                        "#fil_title": self.file_name_value,
                        "%EQUIP1%": wire.label if wire.label else "\u00A0",
                        "%WIRE_MARK%": wire.mark,
                        "%WIRE_SPEC%": wire.wire_spec,
                        "%TERMINAL_UP%": None,
                        "%TERMINAL_DOWN%": None,
                    })

            # 2. Wire Mark Macros
            if wire.mark_pos:
                mx, my = wire.mark_pos
                if not wire.wire_spec or wire.spec_source == "wire":
                    prefix = "wire_mark_no_symbol"
                else:
                    prefix = "wire_mark"
                suffix = "_90" if wire.mark_orientation == "V" else "_0"

                rows.append({
                    "#mac_name": f"{prefix}{suffix}",
                    "#mac_posx": round(mx, 2),
                    "#mac_posy": round(my, 2),
                    "#fil_title": self.file_name_value,
                    "%EQUIP1%": wire.label if wire.label else "\u00A0",
                    "%WIRE_MARK%": wire.mark,
                    "%WIRE_SPEC%": wire.wire_spec,
                    "%TERMINAL_UP%": wire.label,
                    "%TERMINAL_DOWN%": wire.wire_spec if wire.wire_spec else " ",
                })

        # 3. Wire Match Macros
        for cx, cy, word in self.wirematches:
            best_kind = None
            best_d = float("inf")
            for w in self.wires:
                for seg in w.segments:
                    d = pt_seg_dist(cx, cy, seg[0], seg[1])
                    if d < best_d:
                        best_d = d
                        k, _, _ = get_segment_orientation(seg)
                        if k:
                            best_kind = k

            mac_name = "wirematch_90" if best_kind == "V" else "wirematch_0"
            safe_text = "'" + word if word.startswith("=") else word

            rows.append({
                "#mac_name": mac_name,
                "#mac_posx": f"{round(cx, 2)}",
                "#mac_posy": f"{round(cy, 2)}",
                "#fil_title": self.file_name_value,
                "%EQUIP1%": "\u00A0",
                "%WIRE_MARK%": None,
                "%WIRE_SPEC%": None,
                "%TERMINAL_UP%": safe_text,
                "%TERMINAL_DOWN%": "\u00A0",
            })

        return rows

    # ---- Debug visualization ----

    def show_debug_visualization(self, page_label, dxf_extents):
        """Plot wires, labels, and matches over the DXF drawing area."""
        x_min, y_min, x_max, y_max = dxf_extents

        fig = plt.figure(figsize=(16, 12), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_title(f"EXPORT PREVIEW - {page_label} (Close window to continue)")
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        ax.set_aspect('equal')


        # Draw all extracted segments in light gray
        for seg in self.segments:
            (x0, y0), (x1, y1) = seg
            ax.plot([x0, x1], [y0, y1], color='lightgray', linewidth=0.5, alpha=0.5)

        # Draw matched wires in color
        for wire in self.wires:
            color = np.random.rand(3,)
            for (x0, y0), (x1, y1) in wire.segments:
                ax.plot([x0, x1], [y0, y1], c=color, linewidth=2.5, alpha=0.9)
            if wire.mark_pos:
                mx, my = wire.mark_pos
                ax.scatter([mx], [my], color='red', s=40, zorder=5, marker='x')
                tags = []
                if wire.label:
                    tags.append(f"L: {wire.label}")
                if wire.wire_spec:
                    tags.append(f"S: {wire.wire_spec}")
                if wire.wirematch_spec:
                    tags.append(f"M: {wire.wirematch_spec}")
                if tags:
                    label_text = "\n".join(tags)
                    bc = 'green' if wire.spec_source == 'wire' else \
                         'orange' if wire.spec_source == 'cable' else 'blue'
                    ax.text(mx + 0.1, my, label_text, color='white', fontsize=8,
                            weight='bold',
                            bbox=dict(facecolor=bc, alpha=0.8, edgecolor='white',
                                      boxstyle='round,pad=0.2'))

        for cx, cy, word in self.wirematches:
            ax.text(cx, cy, f"Match: {word}", color='yellow', fontsize=7,
                    weight='bold', ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.6, edgecolor='yellow'))

        plt.show()

    # ---- Main processing entry point ----

    def process_dxf(self, plot=False):
        """Process a single DXF file: extract segments, run OCR, match, build rows."""
        print(f"\n{'='*60}")
        print(f"Processing: {self.dxf_path}")
        print(f"{'='*60}")

        self._reset_state()

        # 1. Read DXF
        doc = ezdxf.readfile(self.dxf_path)

        # 2. Extract line segments directly from DXF
        self.extract_segments_from_dxf(doc)

        # 3. Run OCR on the DXF rendering to get text + positions
        ocr_helper = DXFOcrHelper(dpi=self.ocr_dpi)
        ocr_items = ocr_helper.perform_ocr(doc)

        # 4. Classify OCR text into labels, specs, wirematches
        self.extract_labels_from_ocr(ocr_items)

        # 5. Group segments into wires
        self.group_segments()
        print(f"  Grouped into {len(self.wires)} wire groups")

        # 6. Match labels to wires
        self.match_labels_to_wires()
        print(f"  Matched {len(self.matched_labels)} labels to wires")

        # 7. Match spec tags to labels
        self.match_specs_within_radius_of_labels()

        # 8. Filter to only wires with data
        # self.filter_wires()
        # print(f"  {len(self.wires)} wires after filtering")

        # 9. Debug visualization
        if plot:
            if self.segments:
                all_x = [p[0] for seg in self.segments for p in seg]
                all_y = [p[1] for seg in self.segments for p in seg]
                extents = (min(all_x), min(all_y), max(all_x), max(all_y))
            else:
                extents = (0, 0, 1, 1)
            self.show_debug_visualization(Path(self.dxf_path).name, extents)

        # 10. Build output rows
        rows = self.build_rows()
        print(f"  Generated {len(rows)} output rows")
        return rows

    # ---- Excel output ----

    def clear_worksheet_below_header(self, ws, header_row=5):
        if ws.max_row > header_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)

    def write_rows_to_excel_batch(self, rows, sheet_name="Sheet1", header_row=5):
        wb = load_workbook(self.template_xlsx)
        ws = wb[sheet_name]
        self.clear_worksheet_below_header(ws, header_row=header_row)
        start_row = header_row + 1
        for idx, row in enumerate(rows, start=start_row):
            ws.cell(idx, 1, row["#mac_name"])
            ws.cell(idx, 2, row["#mac_posx"])
            ws.cell(idx, 3, row["#mac_posy"])
            ws.cell(idx, 6, row["#fil_title"])
            ws.cell(idx, 21, row["%EQUIP1%"])
            ws.cell(idx, 22, row["%WIRE_MARK%"])
            ws.cell(idx, 24, row["%TERMINAL_UP%"])
            ws.cell(idx, 25, row["%TERMINAL_DOWN%"])
        wb.save(self.template_xlsx)
        print(f"✅ Excel file updated: {self.template_xlsx}")


# -----------------------
# Entry point
# -----------------------

def process_single_file(config, dxf_file, plot=False):
    """Process one DXF file with the given config."""
    cfg = dict(config)
    cfg['dxf_path'] = str(dxf_file)
    cfg['file_name_value'] = Path(dxf_file).stem  # DXF filename without extension
    detector = WireDetectionAndPlacement(**cfg)
    return detector.process_dxf(plot=plot)


if __name__ == "__main__":
    DXF_PATH = "wires_raw/"  # folder of cleaned DXFs, or single .dxf file
    XLSX_PATH = "C:/Users/John.Hoang/Documents/lol_out/wire_template.xlsx"

    # --- DEBUG TOGGLE ---
    DEBUG_MODE = False
    # --------------------

    config = {
        'dxf_path': '',            # set per file
        'seg_min': 0.16,            # inches — minimum segment length to keep
        'connect_dist': 0.005,      # inches — max gap to join segments into a wire
        'label_dist': 0.2,         # inches — max label-to-wire distance
        'template_xlsx': XLSX_PATH,
        'file_name_value': 0,
        'zoom': 2,                 # only for debug viz
        'wire_spec_json': "wire_spec.json",
        'cable_spec_json': "cable_spec.json",
        'ocr_dpi': 600,
    }

    dxf_source = Path(DXF_PATH)

    # Collect DXF files
    if dxf_source.is_dir():
        dxf_files = sorted(dxf_source.glob("*.dxf")) + sorted(dxf_source.glob("*.DXF"))
        print(f"Found {len(dxf_files)} DXF file(s) in {DXF_PATH}")
    else:
        dxf_files = [dxf_source]

    print(f"Starting processing (Debug={DEBUG_MODE})...")
    t0 = time.perf_counter()
    all_results = []

    for i, dxf_file in enumerate(dxf_files):
        if 2<i:
            
            print(f"\n[{i+1}/{len(dxf_files)}] {dxf_file.name}")
            rows = process_single_file(config, dxf_file, plot=DEBUG_MODE)
            all_results.extend(rows)

    print(f"\nProcessing complete. Writing {len(all_results)} rows to Excel...")

    writer = WireDetectionAndPlacement(**{**config, 'dxf_path': str(dxf_files[0])})
    writer.write_rows_to_excel_batch(all_results)

    print(f"Total time: {time.perf_counter() - t0:.2f} seconds")